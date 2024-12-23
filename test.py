import ServerData
import can
import time
import binascii
import sys
import os
import requests
import datetime
from PyQt5.QtGui  import QCursor
from PyQt5.QtCore import QPoint
from openpyxl import Workbook
from datetime import datetime, timedelta,timezone
from PyQt5.QtWidgets import QMainWindow, QApplication,QMessageBox
from PyQt5.QtCore import QDateTime
from PyQt5.QtCore import QTimer
from finalTesting import Ui_FinalTestingUtility
import resources_rc

# Expected CAN IDs and their frame counts
expected_frame_counts = {0x100: 3, 0x101 :3, 0x103 : 3, 0x105 :2, 0x106 :3 , 0x115 : 1, 0x116 : 1,0x109 :1}

# Initialize received_frames with empty lists for each CAN ID
received_frames = {0x100: [],0x101 : [] , 0x103 :[],0x105 :[],0x106 :[] , 0x115 :[], 0x116 : [],0x109:[]}



class MyClass(QMainWindow):
    def __init__(self):
        super().__init__()
        self.ui = Ui_FinalTestingUtility()
        self.ui.setupUi(self)
        self.stackedWidget = self.ui.stackedWidget
        self.bus = None
        self.busy = False
        self.ui.pushButton.clicked.connect(self.goToPage2)
        self.ui.pushButton_8.clicked.connect(self.start_functions)
        self.ui.pushButton_2.clicked.connect(self.save_to_excel)
        self.ui.pushButton.clicked.connect(self.on_button_click)
        self.initialize_can_bus()
        self.IMEI_ascii= None
        self.ICCID_ascii = None
        self.appln_ver = None
        self.GSM_ver = None
        self.Gps_ver = None
        self.Int_vtg = None
        self.mains_vtg = None
        self.Gps_status = None
        self.No_of_Sat = None

        self.stage4_url = "http://192.168.2.253:6101/api/stage4"
        self.stage5_url = "http://192.168.2.253:6101/api/stage5"
        self.device_status_url = "http://192.168.2.253:6101/api/test_points/SENSOR001"
        self.check_server_url = "http://192.168.2.253:6101/api/PROD_check"
        
        self.barcode = "SENSOR001"
        self.model_name = None
        # Initialize a QTimer
        self.timer = QTimer(self)
        self.timer.setInterval(1000)  # 1000 ms = 1 second
        self.timer.timeout.connect(self.on_timer_timeout)
        self.elapsed_time = 0
        self.operator = None
        self.qc_head = None
        # Initialize CAN bus in the __init__ method to avoid reinitializing it every time
        cursor = QCursor()
        cursor.setPos(620, 70)

        self.data = {
            "QR_code": self.barcode,
            "model_name": "ACON4L",
            "final_testing_status": True,
            "IMEI": '123456789',
            "ICCID": "89911234567890123456",
            "SystemRtc": "2024-03-21 15:30:45",
            "AppFWVersion": "1.0.5",
            "BLFWVersion": "2.1.0",
            "GPSFWVersion": "3.2.1",
            "GSMFWVersion": "4.0.2",
            "HWVersion": "V2.0",
            "GPSFix": "3D Fix",
            "HDOP": "1.2",
            "PDOP": "2.4",
            "No_satelite": "8",
            "GSMStatus": "Connected",
            "signalStrength": "-67",
            "Network_code": "40414",
            "Network_Type": "4G",
            "SIM": "Active",
            "MEMS": "Working",
            "Voltage": "12.5",
            "Memory": "75",
            "Ignition": "ON",
            "Tamper": "No",
            "DI_1_H": "1",
            "DI_1_L": "0",
            "DI_2_H": "1",
            "DI_2_L": "0",
            "DI_3_H": "1",
            "DI_3_L": "0",
            "DO_1_H": "1",
            "DO_1_L": "0",
            "DO_2_H": "1",
            "DO_2_L": "0",
            "CAN": "OK",
            "RS485": "Connected",
            "AnalogInput1": "4.5",
            "AnalogInput2": "3.2",
        }
        self.headers = {"Content-Type": "application/json"}

        # Initialize flags
        self.function100_done = False
        self.function101_done = False
        self.function103_done = False
        self.function105_done = False
        self.function106_done = False
        self.function115_done = False
        self.function116_done = False
        self.function109_done = False

        # Timer for delays
        self.timer = QTimer(self)
        self.timer.setSingleShot(True)  # Ensure it only fires once
        self.timer.timeout.connect(self.execute_next_function)

        

    def initialize_can_bus(self):
        try:
            # Initialize the bus once, not inside each function
            self.bus = can.interface.Bus(interface='pcan', channel='PCAN_USBBUS1', bitrate=250000)
            print(f"CAN Bus initialized: {self.bus.channel_info}")
        except can.CanError as e:
            print(f"Error initializing CAN bus: {str(e)}")
            self.bus = None  # Set bus to None if there's an initialization error

    def start_functions(self):
        """Start the series of functions when the button is clicked."""
        print("Button clicked. Starting functions.")
        
        # Reset flags
        self.function100_done = False
        self.function101_done = False
        self.function103_done = False
        self.function105_done = False
        self.function106_done = False
        self.function115_done = False
        self.function116_done = False
        self.function109_done = False
        
        # Call the first function
        self.fun_0x100()

    def fun_0x100(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x100, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x100]):
                message = self.bus.recv(timeout=2)  # 2 second timeout for each frame
                if message:
                    print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x100].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x100]) == expected_frame_counts[0x100]:
                frames = received_frames[0x100]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                print(f"Reassembled message for CAN ID 0x100: {complete_message.hex()}")

                IMEI = complete_message[:15]
                print(f"Extracted IMEI: {IMEI.hex()}")
 
                try:
                  self.IMEI_ascii = IMEI.decode('ascii')  # Decode bytes into ASCII string
                  
                  print(f"Extracted IMEI (ASCII): {self.IMEI_ascii}")
                  self.ui.plainTextEdit_10.setPlainText(self.IMEI_ascii)
                  self.ui.plainTextEdit_12.appendPlainText(f"IMEI : {self.IMEI_ascii}\n")

                  if len(self.IMEI_ascii) < 15:
                     self.ui.plainTextEdit_10.setStyleSheet("background-color: red;")
                  else:
                      self.ui.plainTextEdit_10.setStyleSheet("background-color: white;")
                  
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x100. Expected {expected_frame_counts[0x100]}, but received {len(received_frames[0x100])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x100].clear()
            self.function100_done = True
            time.sleep(2)
            self.execute_next_function()
            print("Frames cleared for CAN ID 0x100")

    def fun_0x101(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x101, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x101]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x101].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x101]) == expected_frame_counts[0x101]:
                frames = received_frames[0x101]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                print(f"Reassembled message for CAN ID 0x101: {complete_message.hex()}")
 
                ICCID = complete_message[:20]
                print(f"Extracted IMEI: {ICCID.hex()}")
 
                try:
                  self.ICCID_ascii = ICCID.decode('ascii')  # Decode bytes into ASCII string
                  print(f"Extracted IMEI (ASCII): {self.ICCID_ascii}")
                  self.ui.plainTextEdit_11.setPlainText(self.ICCID_ascii)
                  self.ui.plainTextEdit_12.appendPlainText(f"ICCID : {self.ICCID_ascii}\n")

                  if len(self.ICCID_ascii)<20:
                      self.ui.plainTextEdit_11.setStyleSheet("background-color: red;")
                  else:
                      self.ui.plainTextEdit_11.setStyleSheet("background-color: white;")
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x101. Expected {expected_frame_counts[0x101]}, but received {len(received_frames[0x101])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x101].clear()
            self.function101_done = True
            time.sleep(2)
            self.execute_next_function()
            print("Frames cleared for CAN ID 0x101")

    def fun_0x103(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x103, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x103]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x103].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x103. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x103]) == expected_frame_counts[0x103]:
                frames = received_frames[0x103]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                print(f"Reassembled message for CAN ID 0x100: {complete_message.hex()}")
 
                try:
                  self.appln_ver = complete_message.decode('ascii')  # Decode bytes into ASCII string
                  print('appln ver ASCII :',self.appln_ver)
                  print(f"Application version: {self.appln_ver}")
                  self.ui.plainTextEdit_8.setPlainText(self.appln_ver)
                  self.ui.plainTextEdit_12.appendPlainText(f"Application Version : {self.appln_ver}\n")

                  if self.appln_ver != 'SAM01_APP_0.0.6_TST06':
                      self.ui.plainTextEdit_8.setStyleSheet("background-color: red;")
                  else:
                      self.ui.plainTextEdit_8.setStyleSheet("background-color: white;")
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x103. Expected {expected_frame_counts[0x103]}, but received {len(received_frames[0x103])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x103].clear()
            self.function103_done = True
            time.sleep(2)
            self.execute_next_function()
            print("Frames cleared for CAN ID 0x103")

    def fun_0x105(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x105, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x105]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x105].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x105]) == expected_frame_counts[0x105]:
                frames = received_frames[0x105]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                print(f"Complete for CAN ID 0x105: {complete_message.hex()}")
 
                try:
                  self.Gps_ver = complete_message.decode('ascii')  # Decode bytes into ASCII string
                  print('GPS ver ASCII :',self.Gps_ver)
                  self.ui.plainTextEdit_5.setPlainText(self.Gps_ver)
                  self.ui.plainTextEdit_12.appendPlainText(f"GPS Version : {self.Gps_ver}\n")
                  #gps_ver_cleaned = self.Gps_ver.strip()
                  print('gps strip',self.Gps_ver)
                  
                  if self.Gps_ver == 'L89HANR01A07S':
                      self.ui.plainTextEdit_5.setStyleSheet("background-color: white;")
                  else:
                      self.ui.plainTextEdit_5.setStyleSheet("background-color: red;")
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x105. Expected {expected_frame_counts[0x105]}, but received {len(received_frames[0x105])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x105].clear()
            self.function105_done = True
            time.sleep(2)
            self.execute_next_function()
            print("Frames cleared for CAN ID 0x105")

    def fun_0x106(self):
        print('fun 0x106 called')
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return
 
        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return
 
        self.busy = True  # Mark the system as busy
        try:
            msg = can.Message(arbitration_id=0x106, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)
           
            # Send the message
            self.bus.send(msg)
            print(f"Message sent on {self.bus.channel_info}")
 
            # Wait for the response
            for i in range(expected_frame_counts[0x106]):
                message = self.bus.recv(timeout=2)  # 1 second timeout for each frame
                if message:
                    print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
                    received_frames[0x106].append(message)
                else:
                    print(f"Timeout waiting for message for CAN ID 0x100. No response received.")
 
            # Check if we have received all expected frames for 0x100
            if len(received_frames[0x106]) == expected_frame_counts[0x106]:
                frames = received_frames[0x106]
                frames.sort(key=lambda x: x.data[0])  # Sort by sequence number
                complete_message = b''.join(frame.data[1:] for frame in frames)
                print(f"Complete message for CAN ID 0x106: {complete_message.hex()}")
 
                try:
                  self.GSM_ver = complete_message.decode('ascii')  # Decode bytes into ASCII string
                  print('GSM ver ASCII :',self.GSM_ver)
                  self.ui.plainTextEdit_6.setPlainText(self.GSM_ver)
                  self.ui.plainTextEdit_12.appendPlainText(f"GSM Version : {self.GSM_ver}\n")

                  if self.GSM_ver != 'EC200UCNAAR03A03M08':
                      self.ui.plainTextEdit_6.setStyleSheet("background-color: red;")
                  else:
                      self.ui.plainTextEdit_6.setStyleSheet("background-color: white;")
                      
                except UnicodeDecodeError:
                  print("Error decoding IMEI to ASCII. The data may contain non-ASCII characters.")
 
            else:
                print(f"Not all frames received for CAN ID 0x106. Expected {expected_frame_counts[0x106]}, but received {len(received_frames[0x106])}.")
 
        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x106].clear()
            self.function106_done = True
            time.sleep(2)
            self.execute_next_function()
            print("Frames cleared for CAN ID 0x106")

    def fun_0x115(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x115, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                exttracted_mains = message.data[1:]
                # Process the received message
                print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
            
                # Decode the received message and update the UI
                self.mains_vtg = exttracted_mains.decode('ascii')  # Decode bytes into ASCII string
                print('Mains vtg:', self.mains_vtg)
            
                # Update the UI with the decoded message
                self.ui.mains_input_2.setPlainText(self.mains_vtg)
                self.ui.plainTextEdit_12.appendPlainText(f"Mains Voltage: {self.mains_vtg}\n")
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x115. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x115].clear()
            self.function115_done = True
            time.sleep(2)
            self.execute_next_function()
            print("Frames cleared for CAN ID 0x115")

    def fun_0x116(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x116, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                exttracted_IntVtg = message.data[1:]
                # Process the received message
                print(f"Received message from CAN ID {hex(message.arbitration_id)}: {message.data.hex()}")
            
                # Decode the received message and update the UI
                self.Int_vtg = exttracted_IntVtg.decode('ascii')  # Decode bytes into ASCII string
                print('Internal vtg:', self.Int_vtg)
            
                # Update the UI with the decoded message
                self.ui.IntBat_input_2.setPlainText(self.Int_vtg)
                self.ui.plainTextEdit_12.appendPlainText(f"Int_Bat Voltage: {self.Int_vtg}\n")
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x116. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x116].clear()
            self.function116_done = True
            time.sleep(2)
            self.execute_next_function()
            print("Frames cleared for CAN ID 0x116")

    def fun_0x109(self):
        if self.busy:  # Check if the system is busy
            print("System is busy, please wait...")
            return

        if self.bus is None:  # Check if the bus was initialized properly
            print("CAN Bus not initialized. Cannot send message.")
            return

        self.busy = True  # Mark the system as busy

        try:
            # Create the CAN message
            msg = can.Message(arbitration_id=0x109, data=[0, 0, 0, 0, 0, 0, 0, 0], is_extended_id=False)

            # Send the message once
            self.bus.send(msg)
            print(f"Message sent on {self.bus.channel_info}")

            # Wait for a response with a timeout (e.g., 2 seconds)
            message = self.bus.recv(timeout=2)  # 2 seconds timeout for response

            if message:
                self.Gps_status = message.data[1]
                print('Gps status :',self.Gps_status)
                self.ui.Operator_2.setPlainText(str(self.Gps_status))
                self.No_of_Sat = message.data[2]
                print('No. of sat:',self.No_of_Sat)
                self.ui.NoOf_satellite.setPlainText(str(self.No_of_Sat))
                self.ui.plainTextEdit_12.appendPlainText(f"GPS status: {str(self.Gps_status)}\n")
                self.ui.plainTextEdit_12.appendPlainText(f"No. of Satellite: {str(self.No_of_Sat)}\n")
            
            else:
                # If no message is received within the timeout period
                print(f"Timeout waiting for message for CAN ID 0x109. No response received.")

        except can.CanError as e:
            print(f"CAN error: {str(e)}")
 
        finally:
            self.busy = False  # Mark the system as not busy
            received_frames[0x109].clear()
            self.function109_done = True
            time.sleep(2)
            self.execute_next_function()
            print("Frames cleared for CAN ID 0x109")


    


    def execute_next_function(self):
        """Check which function is done and call the next one."""
        if self.function100_done and not self.function101_done:
            self.fun_0x101()  # Call function 2 after function 1 is done

        elif self.function101_done and not self.function103_done:
             self.fun_0x103()  # Call function 3 after function 2 is done

        elif self.function103_done and not self.function105_done:
            self.fun_0x105()

        elif self.function105_done and not self.function106_done:
            self.fun_0x106()

        elif self.function106_done and not self.function115_done:
            self.fun_0x115()

        elif self.function115_done and not self.function116_done:
            self.fun_0x116()

        elif self.function116_done and not self.function109_done:
            self.fun_0x109()
        else:
            print("All functions completed.")
            # You can enable a button or perform other tasks once all functions are done
            self.ui.pushButton_2.setEnabled(True)  # Enable button after all functions are done

 
 
    def login(self):
        # Get the username and password entered by the user
        Operator = self.ui.plainTextEdit.toPlainText()
        QC_head = self.ui.plainTextEdit_2.toPlainText()
 
        # Logic to check the username and password
        if Operator is not None and QC_head is not None:
            #self.ui.pushButton.clicked.connect(self.goToPage2)
            self.show_message("Login Successful", "Welcome, Operator!")
        else:
            self.show_message("Login Failed", "Invalid username or password. Please try again.")
 
    def goToPage2(self):
        self.stackedWidget.setCurrentIndex((self.stackedWidget.currentIndex()+1)%2)

    def clean_string(self,input_string):
    
        # Filter out characters with ASCII values from 0-31 and 127
        return ''.join(c for c in input_string if 31 < ord(c) < 127)

    def save_to_excel(self):
    
        # Create a new workbook and select the active sheet
        wb = Workbook()
        ws = wb.active

        # Set the headers for the columns
        ws['A1'] = 'Operator'
        ws['B1'] = 'QC Head'
        ws['C1'] = 'IMEI'
        ws['D1'] = 'ICCID'
        ws['E1'] = 'Application Version'
        ws['F1'] = 'GSM Version'
        ws['G1'] = 'GPS Version'
        ws['H1'] = 'Mains vtg'
        ws['I1'] = 'Int_Bat vtg'
        ws['J1'] = 'GPS status'
        ws['K1'] = 'No.of Sat'

        # Clean the data before inserting into the worksheet
        ws['A2'] = self.clean_string(self.operator)
        ws['B2'] = self.clean_string(self.qc_head)
        ws['C2'] = self.clean_string(self.IMEI_ascii)
        ws['D2'] = self.clean_string(self.ICCID_ascii)
        ws['E2'] = self.clean_string(self.appln_ver)
        ws['F2'] = self.clean_string(self.GSM_ver) 
        ws['G2'] = self.clean_string(self.Gps_ver)
        ws['H2'] = self.clean_string(self.mains_vtg)
        ws['I2'] = self.clean_string(self.Int_vtg)
        ws['J2'] = self.clean_string(str(self.Gps_status))
        ws['K2'] = self.clean_string(str(self.No_of_Sat))

        # Get the current working directory
        current_directory = os.getcwd()

        # Set the path for the Excel file
        file_path = os.path.join(current_directory, "saved_data.xlsx")

        try:
            # Save the workbook to the specified path
            wb.save(file_path)
            print(f"Data saved to {file_path}")
            QMessageBox.information(self, "Success", f"Data successfully saved to {file_path}")

        except Exception as e:
            print(f"Error saving data to Excel: {str(e)}")
            QMessageBox.warning(self, "Error", f"Failed to save data to Excel: {str(e)}")

    def check_barcode(self):
        # Get the barcode text from the input field
        self.ui.barcode_Input_2.setFocus()  # Set focus to the barcode input field
        time.sleep(0.1)
        #self.barcode = self.ui.barcode_Input_2.toPlainText().strip()  # Get text and remove any leading/trailing whitespaces
        #self.barcode = self.barcode.replace('\n', '').replace('\r', '').strip()
        print(f"Raw Barcode: {repr(self.barcode)}")
        print(f"Barcode: {self.barcode}")

        # Check if barcode text is valid (non-empty)
        if self.barcode:
            self.ui.barcode_Input_2.setPlainText(str(self.barcode))  # Copy the barcode to another input field

            # Barcode is received, make the API request
            # try:
            #     response = requests.put(self.stage4_url, json=self.data, headers=self.headers)
            #     print(response)
            #
            #     if response.status_code == 200:
            #         print('Barcode data scanned successfully')
            #         self.ui.pushButton_2.setEnabled(True)  # Enable the button if barcode data is valid
            #         self.check_previous_stages()  # Continue with previous stages
            #         self.ui.barcode_Input_2.setPlainText(self.barcode)  # Set actual barcode value
            #     else:
            #         print(f'Error reading barcode: {response.status_code}')
            #         self.ui.barcode_Input_2.setPlainText(
            #             'Error reading barcode')  # Set error message if barcode scan fails
            # except requests.exceptions.RequestException as e:
            #     print(f'API request failed: {e}')
            #     self.ui.barcode_Input_2.setPlainText('API request failed')  # Set API request failure message

    def on_button_click(self):
        # Call both actions (goToPage2 and login) in sequence
        print("Button clicked")

        self.login()
        self.showDateTime()
        self.check_server_status()
        self.check_barcode()
        if not self.timer.isActive():
            self.timer.start()  # Start the timer if not already active

    def showDateTime(self):
        # Get the current date and time
        current_datetime = QDateTime.currentDateTime().toString()

        # Display the date and time in the QPlainTextEdit widget
        self.ui.operator_Input_2.setPlainText(current_datetime)

    def show_message(self, title, message):
        # Show a message box to the user
        msg_box = QMessageBox()
        msg_box.setWindowTitle(title)
        msg_box.setText(message)
        msg_box.exec_()

    def on_timer_timeout(self):
        # This method is called every time the timer times out (every second)
        print(f"Timer triggered. Elapsed time: {self.elapsed_time} seconds")

        # Increment the elapsed time by 1 second
        self.elapsed_time += 1

        hours, remainder = divmod(self.elapsed_time, 3600)  # Get hours and remainder
        minutes, seconds = divmod(remainder, 60)  # Get minutes and seconds

        # Format the time as hh:mm:ss
        formatted_time = f"{hours:02}:{minutes:02}:{seconds:02}"

        # Update the operator_Input_3 field with the elapsed time
        self.ui.operator_Input_3.setPlainText(f"{formatted_time}")

    def login(self):
        # Retrieve the text from the text fields
        self.operator = self.ui.plainTextEdit.toPlainText()  # Retrieve the operator input
        self.qc_head = self.ui.plainTextEdit_2.toPlainText()  # Retrieve the QC head input

        # Debugging print statements to check the retrieved values
        print("operator:", repr(self.operator))  # Using repr() to detect empty strings
        print("qc_head:", repr(self.qc_head))

        # Check if both fields are filled
        if self.operator == "" or self.qc_head == "":
            print("Either operator or qc_head is empty!")  # Debugging message
            self.show_message("Login Failed", "Please Enter operator and qc head")
            # Ensure navigation doesn't happen when either field is empty
        else:
            # If both fields are filled, proceed with the login success logic
            print("operator:", self.operator)
            print("qc_head:", self.qc_head)

            self.get_device_model()
            #self.show_message("Login Successful", "Welcome, Operator!")

            # Set the values in the corresponding UI input fields
            self.ui.operator_Input.setPlainText(self.operator)
            self.ui.QC_input.setPlainText(self.qc_head)

            # Navigate to Page 2 only if both fields are filled
            self.stackedWidget.setCurrentIndex(1)  # Go to Page 2

    def get_device_model(self):
        # Make the GET request to fetch device information based on barcode
        print("get_device_model")
        # Include the barcode in the payload (or query if required by the API)
        try:
            # Make the API request
            response = requests.put(self.stage5_url, json=self.data, headers=self.headers)
            print("get_device_model",response)
            if response.status_code == 200:
                # Parse the response JSON
                response_data = response.json()
                self.ui.barcode_Input_2.setPlainText(self.barcode)
                # Extract the model name from the response
                self.model_name = response_data.get("device", {}).get("model_name", "Model name not found")
                self.ui.QC_input_2.setPlainText(self.model_name)
                if self.model_name is not None:
                    self.select_parameters()
                # Display the model name in the UI or print it
                    print(f"Device Model: {self.model_name}")

                # If you want to show it in the UI, you can use something like:
                # self.ui.modelNameLabel.setText(model_name)

            else:
                print(f"Error fetching device data. Status code: {response.status_code}")

        except requests.exceptions.RequestException as e:
            # Handle any errors in the request
            print(f"Error: {e}")

    def goToPage2(self):
        print("Navigating to Page 2")
        self.stackedWidget.setCurrentIndex((self.stackedWidget.currentIndex() + 1) % 2)

    def check_server_status(self):
        response = requests.get("http://192.168.2.253:6101/api/PROD_check")
        if response.status_code == 200:
            print("response",response.status_code)
            self.ui.barcode_Input.setPlainText("Connected")
        else:
            self.ui.barcode_Input.setPlainText("Please check connection")

    def check_previous_stages(self):
        url = "http://192.168.2.253:6101/api/test_points/SENSOR001"

        try:
            # Send a GET request to the URL
            response = requests.get(url)

            # Check if the request was successful (status code 200)
            if response.status_code == 200:
                data = response.json()  # Parse the JSON response

                # Access the mechanical_fitting_status
                mechanical_fitting_status = data.get('device', {}).get('status', {}).get('mechanical_fitting_status',
                                                                                         False)
                print("mechanical_fitting_status", mechanical_fitting_status)
                # Check the status
                if mechanical_fitting_status:
                    print("Previous stage passed!")
                else:
                    print("Previous stage not passed.")
            else:
                print(f"Error: Unable to retrieve data, status code {response.status_code}")

        except requests.exceptions.RequestException as e:
            print(f"Request failed: {e}")

    def select_parameters(self):
        print("select_parameters", self.model_name)
        if self.model_name == "ACON4L":

            self.ui.DI1_H_6.hide()
            self.ui.DI1_H_7.hide()
            self.ui.DI1_H_4.hide()
            self.ui.DI_H.hide()
            self.ui.DI1_H_5.hide()
            self.ui.DI1_H_8.hide()
            self.ui.DO1_L.hide()
            self.ui.DO1_H.hide()
            self.ui.DO2_L.hide()
            self.ui.DO2_H.hide()


            self.ui.label_50.hide()
            self.ui.label_49.hide()
            self.ui.label_40.hide()
            self.ui.label_41.hide()
            self.ui.label_41.hide()
            self.ui.label_42.hide()
        elif self.model_name == "Sampark AIS":
            self.ui.DI1_H_6.setDisabled(False)
            self.ui.DI1_H_7.setDisabled(False)
            self.ui.DI1_H_4.setDisabled(False)
            self.ui.DI_H.setDisabled(False)
            self.ui.DI1_H_5.setDisabled(False)
            self.ui.DI1_H_8.setDisabled(False)
            self.ui.DO1_L.setDisabled(False)
            self.ui.DO1_H.setDisabled(False)
            self.ui.DO2_L.setDisabled(False)
            self.ui.DO2_H.setDisabled(False)

    
    
    def send_data(self):
        print('button clicked')
        myobj = {"Status":True,
            "QR_code": "SENSOR965",  # This can be dynamically changed
    "visual_inspection_status": True,
    "visual_inspection_timestamp": {
        "$date": datetime.now(timezone.utc).isoformat()
    },
    "board_flashing_status": True,
    "board_flashing_timestamp": {
        "$date": datetime.now(timezone.utc).isoformat()
    },
    "screw_fitting_status": None,
    "mechanical_fitting_status": None,
    "mechanical_fitting_timestamp": None,
    "final_testing_status": None,
    "final_testing_timestamp": None,
    "IMEI": self.IMEI_ascii,
    "ICCID": "78910",
    "SystemRtc": None,
    "AppFWVersion": "88888",
    "BLFWVersion": None,
    "GPSFWVersion": None,
    "GSMFWVersion": None,
    "HWVersion": None,
    "GPSFix": None,
    "HDOP": None,
    "PDOP": 'xxxxx',
    "No_satelite": None,
    "GSMStatus": 'yyyyy',
    "signalStrength": None,
    "Network_code": None,
    "Network_Type": None,
    "SIM": None,
    "MEMS": None,
    "Voltage": None,
    "Memory": '111111',
    "Ignition": None,
    "Tamper": None,
    "DI_1_H": None,
    "DI_1_L": None,
    "DI_2_H": None,
    "DI_2_L": None,
    "DI_3_H": None,
    "DI_3_L": None,
    "DO_1_H": None,
    "DO_1_L": None,
    "DO_2_H": None,
    "DO_2_L": '22222',
    "CAN": None,
    "RS485": None,
    "AnalogInput1": None,
    "AnalogInput2": None,
    "sticker_printing_status": None,
    "sticker_printing_timestamp": None,
    "last_updated": {
        "$date": datetime.now(timezone.utc).isoformat()
    },
    "UID": None
        }
        headers = {'Content-Type': 'application/json'}  # Ensure correct headers
        x = requests.put(self.stage5_url, json=myobj, headers=headers,)
        print(x.status_code)

        if x.status_code == 200:
            print("POST request successful!")
            print(x.text)  # Print the response from the server
        else:
            print(f"POST request failed with status code {x.status_code}: {x.text}")


       
        

   
    

   
    

# Entry point of the program
if __name__ == "__main__":
    app = QApplication(sys.argv)    
    # Create an instance of the MyClass class
    processor = MyClass()
    
    processor.show()
    sys.exit(app.exec_())
